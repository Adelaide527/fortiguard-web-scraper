from sys import flags
from bs4 import BeautifulSoup
import time
from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By 
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoSuchWindowException
from openpyxl import Workbook

global i #Oh god. I was told to never do this. but I need a global variable.
global file_name

file_name="fortiguard_threats_2.xlsx"

#Prep the browser
service = Service('C:/WebDriver/bin/geckodriver.exe') #Define the service. geckodriver=Firefox
browser = webdriver.Firefox(service=service) #Using FireFox browser

#Create file to save everything into
# outfile = open("output.txt", 'w')
wb = Workbook()
wb.create_sheet("Page1") #Create our first sheet
ws = wb["Page1"] #Write to it
i = 1 #Our page number

class Threat:
  #####
  # Variable list:
  # name, description, affected_products, impact, reccomended_actions
  #####
  def __init__(self):
    self.name = ""
    self.description = ""
    self.affected_products = ""
    self.impact = ""
    self.reccomended_actions = ""

  def set_name(self, name):
    self.name = name
  
  def set_description(self, description):
    self.description = description

  def set_affected_products(self, affected_products):
    self.affected_products = affected_products

  def set_impact(self, impact):
    self.impact = impact

  def set_reccomended_actions(self, reccomended_actions):
    self.reccomended_actions = reccomended_actions
  
  def print(self):
    return str(self.name + "\n")
    # return str(self.name + " | " + self.description + " | " + self.affected_products + " | " + self.impact + " | " + self.reccomended_actions)

  def write_to_ws(self, worksheet, row):
    print(row)
    worksheet.cell(row=row, column=1, value=self.name)
    worksheet.cell(row=row, column=2, value=self.description)
    worksheet.cell(row=row, column=3, value=self.affected_products)
    worksheet.cell(row=row, column=4, value=self.impact)
    worksheet.cell(row=row, column=5, value=self.reccomended_actions)
    
browser.get('https://www.fortiguard.com/encyclopedia?type=ips&page=101') #Load page
time.sleep(3) #Wait for page to load (giving 3 seconds)

def get_details(browser, threat, i, link):
  try:
    threat.set_name(browser.find_element(By.CLASS_NAME, "detail-item").text)
    threat.set_description(browser.find_element(By.XPATH, "//div[@class='detail-item'][2]/p").text)
    threat.set_affected_products(browser.find_element(By.XPATH, "//div[@class='detail-item'][3]/p").text)
    if(browser.find_element(By.XPATH, "//div[@class='detail-item'][4]/p").text): # If we can find it
      threat.set_impact(browser.find_element(By.XPATH, "//div[@class='detail-item'][4]/p").text) # put
    else: #We put N/A and I will FIX IT LATER
      threat.set_impact("N/A")
    if(browser.find_element(By.XPATH, "//div[@class='detail-item'][5]/p").text): # If we can find it
      threat.set_reccomended_actions(browser.find_element(By.XPATH, "//div[@class='detail-item'][5]/p").text)
    else: #We put N/A and I will FIX IT LATER
      threat.set_reccomended_actions("N/A")

    # grab the active worksheet
    worksheet = wb.active
    threat.write_to_ws(worksheet, i)
    # outfile.write(threat.print()) #Write to our output file
  except NoSuchElementException:
    time.sleep(5) #We will give it 5 more seconds to load the page
    try:
      get_details(browser, threat, i, link) #Then try again
    except NoSuchElementException: #If we still can't find it
      browser.execute_script(f"location.href='{link}';") #Fetch the link again
      time.sleep(15) #Give it 10 seconds to laod in
      get_details(browser, threat, i, link) #And try again

def page_scrape(browser, i):
  elements = browser.find_elements(By.CLASS_NAME, "title") #Find all titles
  original_window = browser.current_window_handle #Store the ID of the original window
  assert len(browser.window_handles) == 1 #Check we don't have other windows open already
  
  for element in elements:
    time.sleep(3) #Wait for page to load (giving 3 seconds)
    link = element.find_element(By.LINK_TEXT, element.text) #Find where the link is
    browser.execute_script("window.scrollBy(0,300)") #Scroll to avoid not in scope error
    link = link.get_attribute("href") #Get link
    browser.switch_to.new_window('window') #Switch to new window
    
    time.sleep(4) #Wait for page to load (giving 4 seconds)
    browser.execute_script(f"location.href='{link}';") #Go to the link we saved
    time.sleep(4) #Wait for page to load (giving 4 seconds)
    threat = Threat() #Define our threat (create instance of Threat class)
    
    #Saving the elements of our threat, as defined in our class object
    #We are looking for them by their class name "detail-item"
    #Then, we are going to their child element (as necessary) so that we will only be saving the paragraph(s)
    try:
      get_details(browser, threat, i, link)
      
      #If we cannot read the threats
    except NoSuchElementException:
      time.sleep(5) #We will give it 5 more seconds to load the page
      try:
        get_details(browser, threat, i, link) #Then try again
      except NoSuchElementException: #If we still can't find it
        browser.execute_script(f"location.href='{link}';") #Fetch the link again
        time.sleep(15) #Give it 10 seconds to laod in
        get_details(browser, threat, i, link) #And try again
    except NoSuchWindowException:
      browser.refresh() #We will try refreshing the page
      time.sleep(5) #Give it 5 seconds to load in 
      get_details(browser, threat, i, link) #And try again
    
    browser.close() #Close the current tab
    i=i+1 #Can't use ++ in python? Ew. But we're iterating to the next row in our worksheet.
    browser.switch_to.window(original_window) #Switch back to the original window


page_scrape(browser, i) #Scrape first page
while(browser.find_element(By.XPATH, "//*[@aria-label='Next']")): #While we still have a next page
  i=i+20
  wb.save(file_name)
  browser.find_element(By.XPATH, "//*[@aria-label='Next']").click() #We click on it
  if(i<2000):
    page_scrape(browser, i) #Scrape the new page.
  else: #If we reach page 100
    wb.save(file_name) #We will do our final save
    browser.quit() #Done with Browser